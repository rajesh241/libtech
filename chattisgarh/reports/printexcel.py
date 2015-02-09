def querytoexcel(infilepath,infilename,inquery,indbname):
  import csv
  #from bs4 import BeautifulSoup
  import requests
  import MySQLdb
  import os
  import time
  import re
  import sys
  from openpyxl import Workbook
  from openpyxl.compat import range
  from openpyxl.cell import get_column_letter
  
  wb = Workbook()
  dest_filename =  infilepath+infilename
  ws = wb.active
  ws.title = "range names"
  
  
  #Connect to MySQL Database
  db = MySQLdb.connect(host="localhost", user="root", passwd="golani123", db="surguja",charset='utf8')
  cur=db.cursor()
  db.autocommit(True)
  #Query to set up Database to read Hindi Characters
  query="SET NAMES utf8"
  cur.execute(query)
  query="use "+indbname  
  cur.execute(query)
  query=inquery
  #print query
  cur.execute(query)
  results = cur.fetchall()
  num_fields = len(cur.description)
  #print num_fields
  num_rows=cur.rowcount
  #print num_rows
  i=1
  col_idx=0
  for colheader in cur.description:
    col_idx=col_idx+1
    j = get_column_letter(col_idx)
    ws.cell('%s%s'%(j, i)).value = '%s' % (colheader[0])
  for row in results:
    i=i+1
    for col_idx in range(1, num_fields+1):
      j = get_column_letter(col_idx)
      ws.cell('%s%s'%(j, i)).value = '%s' % (row[col_idx-1])
    #print row[0]
  ws = wb.create_sheet()
  ws.title = 'Pi'
  ws['F5'] = 3.14
  wb.save(filename = dest_filename)
